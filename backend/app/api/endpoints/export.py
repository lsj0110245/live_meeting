from typing import Any
import re
import os
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.api import deps
from app.db.session import get_db
from app.models.meeting import Meeting
from app.models.transcript import Transcript
from app.models.summary import Summary
from app.models.user import User
import io
import pandas as pd
from datetime import datetime
import openpyxl
from urllib.parse import quote

router = APIRouter()


def parse_ai_summary(summary_content: str) -> dict:
    """
    AI 요약 내용을 섹션별로 파싱 (요약, 주요 안건, 결정 사항, 향후 계획)
    """
    result = {
        'summary': '',
        'agenda_content': '',
        'decisions': '',
        'actions': ''
    }
    
    try:
        # 1. 📅 요약 섹션
        summary_match = re.search(r'##\s*📅\s*요약\s*\n(.*?)(?=##|$)', summary_content, re.DOTALL)
        if summary_match:
            result['summary'] = summary_match.group(1).strip()
        
        # 2. 📌 주요 안건 및 내용
        # 팁: '주요 안건'만 있는 경우도 대비하여 유연하게 매칭
        agenda_match = re.search(r'##\s*📌\s*주요 안건(?: 및 내용)?\s*\n(.*?)(?=##|$)', summary_content, re.DOTALL)
        # 만약 별도의 '상세 논의 내용' 섹션이 있다면 (이전 버전 호환성)
        discussion_match = re.search(r'##\s*💬\s*상세 논의 내용\s*\n(.*?)(?=##|$)', summary_content, re.DOTALL)
        
        agenda_parts = []
        if agenda_match:
            agenda_parts.append(agenda_match.group(1).strip())
        if discussion_match:
            agenda_parts.append("【상세 논의】\n" + discussion_match.group(1).strip())
            
        result['agenda_content'] = '\n\n'.join(agenda_parts) if agenda_parts else ''
        
        # 3. ✅ 결론 및 결정 사항
        # 팁: '결정 사항'만 있는 경우도 대비
        decision_match = re.search(r'##\s*✅\s*(?:결론 및 )?결정 사항\s*\n(.*?)(?=##|$)', summary_content, re.DOTALL)
        if decision_match:
            result['decisions'] = decision_match.group(1).strip()
            
        # 4. 📝 향후 계획
        action_match = re.search(r'##\s*📝\s*향후 계획.*?\n(.*?)(?=##|$)', summary_content, re.DOTALL)
        if action_match:
            result['actions'] = action_match.group(1).strip()
            
        # [추가] 엑셀에 불필요한 제목 문구 제거 (사용자 요청)
        # LLM이 내용 앞에 "📅 **요약**" 등을 반복해서 넣는 경우 이를 제거합니다.
        for key in result:
            if result[key]:
                # 이모지 + **제목** + 줄바꿈/공백 패턴 제거
                result[key] = re.sub(r'^(?:📅|📌|✅|📝|💬)\s*\*\*.*?\*\*\s*\n*', '', result[key]).strip()
                # 혹시 제목이 없는 **내용** 형태도 시작부분에 있으면 제거
                result[key] = re.sub(r'^\*\*.*?\*\*\s*\n*', '', result[key]).strip()
        
    except Exception as e:
        print(f"AI 요약 파싱 오류: {str(e)}")
        result['summary'] = summary_content
    
    return result


@router.get("/{meeting_id}", response_class=StreamingResponse)
def export_meeting(
    meeting_id: int,
    format: str = "csv",  # csv or xlsx
    db: Session = Depends(get_db),
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    """
    회의록 내보내기 (CSV / Excel)
    - CSV: 전사 내용만 (한글 인코딩 수정)
    - Excel: source.xlsx 템플릿 기반 (메타데이터 + AI 요약)
    """
    # 1. 회의 권한 확인
    meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
    if not meeting:
        raise HTTPException(status_code=404, detail="회의를 찾을 수 없습니다.")
    if meeting.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="권한이 없습니다.")
        
    # 2. 데이터 조회
    transcripts = db.query(Transcript).filter(Transcript.meeting_id == meeting_id).order_by(Transcript.start_time).all()
    summary = db.query(Summary).filter(Summary.meeting_id == meeting_id).first()
    
    # 3. 파일 생성 및 반환
    filename = f"{meeting.title}_{datetime.now().strftime('%Y%m%d')}"
    
    if format == "csv":
        # CSV: Excel과 동일한 형식 (메타데이터 + AI 요약)
        # AI 요약 파싱
        parsed_summary = {'summary': '', 'agenda_content': '', 'decisions': '', 'actions': ''}
        if summary:
            parsed_summary = parse_ai_summary(summary.content)
        
        # 데이터 구성 (10개 항목, 회의 목적은 수동 입력값 meeting.purpose 사용)
        data = [
            ["회의명", meeting.title or ''],
            ["회의 유형", meeting.meeting_type or ''],
            ["회의일시", meeting.meeting_date.strftime('%Y-%m-%d %H:%M') if meeting.meeting_date else ''],
            ["참석자", meeting.attendees or ''],
            ["작성자", meeting.writer or ''],
            ["회의 목적", meeting.description or ''],
            ["요약", parsed_summary['summary']],
            ["주요 안건 및 내용", parsed_summary['agenda_content']],
            ["결론 및 결정 사항", parsed_summary['decisions']],
            ["향후 계획", parsed_summary['actions']]
        ]
        
        df = pd.DataFrame(data, columns=["항목", "내용"])
        
        stream = io.StringIO()
        # 한글 인코딩 수정: utf-8-sig (BOM 추가)
        df.to_csv(stream, index=False, encoding='utf-8-sig')
        response = StreamingResponse(
            iter([stream.getvalue().encode('utf-8-sig')]), 
            media_type="text/csv; charset=utf-8"
        )
        encoded_filename = quote(filename)
        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}.csv"
        return response
        
    elif format == "xlsx":
        # Excel: source.xlsx 템플릿 사용
        template_path = os.path.join(os.getcwd(), "source.xlsx")
        
        # AI 요약 파싱
        parsed_summary = {'summary': '', 'agenda_content': '', 'decisions': '', 'actions': ''}
        if summary:
            parsed_summary = parse_ai_summary(summary.content)
            
        if os.path.exists(template_path):
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # 데이터 매핑 (회의 목적은 수동 데이터 meeting.purpose 사용)
            # 데이터 매핑 (B열은 제목, C열에 내용 입력 / 2행부터 시작)
            ws['C2'] = meeting.title or ''
            ws['C3'] = meeting.meeting_type or ''
            ws['C4'] = meeting.meeting_date.strftime('%Y-%m-%d %H:%M') if meeting.meeting_date else ''
            ws['C5'] = meeting.attendees or ''
            ws['C6'] = meeting.writer or ''
            ws['C7'] = meeting.description or ''
            ws['C8'] = parsed_summary['summary']
            ws['C9'] = parsed_summary['agenda_content']
            ws['C10'] = parsed_summary['decisions']
            ws['C11'] = parsed_summary['actions']
            
            # 스타일 설정 (C열 2~11행에 줄바꿈 허용 및 자동 행 높이 조절)
            from openpyxl.styles import Alignment
            for row in range(2, 12):
                cell = ws.cell(row=row, column=3) # C열
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                # 긴 내용 행 높이 자동 조절 (목적 ~ 향후 계획: 7~11행)
                if row >= 7:
                    ws.row_dimensions[row].height = None
            
            # 열 너비 조정 (B열: 제목, C열: 내용)
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 80
        else:
            # 템플릿 없을 경우 Fallback (코드로 생성)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "회의록"
            
            from openpyxl.styles import Font, Alignment, PatternFill
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            rows_data = [
                ("회의명", meeting.title or ''),
                ("회의 유형", meeting.meeting_type or ''),
                ("회의일시", meeting.meeting_date.strftime('%Y-%m-%d %H:%M') if meeting.meeting_date else ''),
                ("참석자", meeting.attendees or ''),
                ("작성자", meeting.writer or ''),
                ("회의 목적", meeting.description or ''),
                ("요약", parsed_summary['summary']),
                ("주요 안건 및 내용", parsed_summary['agenda_content']),
                ("결론 및 결정 사항", parsed_summary['decisions']),
                ("향후 계획", parsed_summary['actions'])
            ]
            
            for idx, (label, value) in enumerate(rows_data, start=2): # 2행부터 시작
                # B열: 레이블
                cell_b = ws.cell(row=idx, column=2, value=label)
                cell_b.font = header_font
                cell_b.fill = header_fill
                cell_b.alignment = Alignment(horizontal='left', vertical='top')
                
                # C열: 값
                cell_c = ws.cell(row=idx, column=3, value=value)
                cell_c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 80
            
            for idx in range(7, 12): # 목적 ~ 향후 계획 행 높이 자동조절
                ws.row_dimensions[idx].height = None
        
        # BytesIO로 저장
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        response = StreamingResponse(
            iter([stream.getvalue()]), 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        encoded_filename = quote(filename)
        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}.xlsx"
        return response
        
    else:
        raise HTTPException(status_code=400, detail="지원하지 않는 형식입니다. (csv, xlsx)")
